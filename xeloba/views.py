from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.template import loader
from django.middleware.csrf import get_token
import json
from collections import Counter
from datetime import datetime, time
import time as smstimer
import os
import requests
import openpyxl
import random
from .models import Adminlogin, Craftsman
from plyer import notification



# import sched
# import schedule
# random generator and time date

random_number = random.randint(100000, 900000)
rn = f'user_{random_number}'
current_datetime = datetime.now()
formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
dateshorts = current_datetime.strftime('%d-%m-%Y')
# end

print(formatted_datetime)

api_url = 'http://127.0.0.1:5500/data.json'
response = requests.get(api_url)

if response.status_code == 200:
    data = response.json()
else:
    data = []

categories = ''



getBalance = 'https://smsoffice.ge/api/getBalance/?key=f5e53ecd65894293934d40e5c3bcd901'

balance = requests.get(getBalance)

bl = ''

if balance.status_code == 200:
    try:
        # Parse the JSON response
        balance_data = balance.json()
        
        bl = balance_data
    except ValueError:
        print("Error: Unable to parse JSON response.")
else:
    print(f"Error: Request failed with status code {response.status_code}")

print(balance)

def admlogin(request):
    c = {'csrf_token': get_token(request)} 
    template = loader.get_template('admlogin.html')

    user_data = {
    'token': c,
    
    }

    # Define the file path for the JSON file
    json_file_path = './token.json'  # Change this to your desired file path

    # Write the data to the JSON file
    with open(json_file_path, 'w') as json_file:
        json.dump(user_data, json_file,indent=4)

    return HttpResponse(template.render(c, request)) 
    


def login(request):
    if request.method == "POST":
        user = request.POST.get('username')
        password = request.POST.get('password')

        userdb = Adminlogin.objects.all()

        for us in userdb:
            if us.username == user and us.password == password:
                
                newUserNotify()

                api_url = 'http://127.0.0.1:5500/data.json'
                response = requests.get(api_url)

                if response.status_code == 200:
                    data = response.json()
                    categories = set(user['profession'] for user in data)

                    return render(request, 'adminpanel.html', {'users': data, 'categories': categories, 'smsbalance': bl})
                else:
                    return HttpResponse("Failed to fetch data from API", status=response.status_code)
        
        return HttpResponse("Invalid username or password", status=401)
    else:
        return HttpResponse("Method not allowed", status=405)

def statistics():
    api_url = 'http://127.0.0.1:5500/data.json'
    response = requests.get(api_url)

    if response.status_code == 200:
        data = response.json()

        # Print the fetched data for debugging
        print("Fetched data:", data)

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()

        # Select the active worksheet and change its name
        sheet1 = workbook.active
        sheet1.title = "User Data"

        # Add headers to the first worksheet
        headers = ['UserID', 'Set Date', 'Name', 'Lastname', 'Personal ID', 'Email', 'Phone Number', 'Password', 'City', 'District', 'Status', 'Profession', 'Experience', 'Price', 'Profile Image']
        for col_num, header in enumerate(headers, start=1):
            sheet1.cell(row=1, column=col_num, value=header)

        # Populate data from the API response into the first sheet
        for i, user_data in enumerate(data, start=2):
            sheet1[f'A{i}'] = user_data.get('userID', '')
            sheet1[f'B{i}'] = user_data.get('setDate', '')
            sheet1[f'C{i}'] = user_data.get('name', '')
            sheet1[f'D{i}'] = user_data.get('lastname', '')
            sheet1[f'E{i}'] = user_data.get('personalID', '')
            sheet1[f'F{i}'] = user_data.get('email', '')
            sheet1[f'G{i}'] = user_data.get('phoneNumber', '')
            sheet1[f'H{i}'] = user_data.get('password', '')
            sheet1[f'I{i}'] = user_data.get('city', '')
            sheet1[f'J{i}'] = user_data.get('district', '')
            sheet1[f'K{i}'] = user_data.get('status', '')
            sheet1[f'L{i}'] = user_data.get('profession', '')
            sheet1[f'M{i}'] = user_data.get('experience', '')
            sheet1[f'N{i}'] = user_data.get('price', '')
            sheet1[f'O{i}'] = user_data.get('profileimgs', '')

        # Create a second worksheet and name it
        sheet2 = workbook.create_sheet(title="sms")

        # Add headers to the second worksheet
        additional_headers = ['chack date', 'sms']
        for col_num, header in enumerate(additional_headers, start=1):
            sheet2.cell(row=1, column=col_num, value=header)

        # Populate data from the API response into the second sheet
        sheet2['A2'] = formatted_datetime
        sheet2['B2'] = bl
    

        # Save the workbook
        workbook.save(f'reports/userReport-{dateshorts}.xlsx')

        # Print a message indicating successful saving
        return "Excel file created successfully"

    else:
        # Print an error message if failed to fetch data from API
        return "Failed to fetch data from API"

statistics()


def newUserNotify():
    api_url = 'http://127.0.0.1:5500/data.json'
    response = requests.get(api_url)
    
    # Check if the request was successful
    if response.status_code == 200:
        try:
            data = response.json()
        except ValueError as e:
            print(f'Error parsing JSON: {e}')
            return
        
        # Debugging: Print the parsed JSON data
        print(f'Parsed JSON Data: {data}')
        
        
        # Adjust the path to your local image file
        image_path = os.path.abspath(os.path.join('static', 'image', 'xelobalogo.jpeg'))
        
        # Debugging: Print the current date and image path
        print(f'Current Date: {dateshorts}')
        print(f'Image Path: {image_path}')
        
        new_users = [user for user in data if user['setDate'].startswith(dateshorts)]
        
        # Debugging: Print the list of new users found
        print(f'New Users: {new_users}')
        lng = len(new_users)
        if new_users:
            for user in new_users:
                notification.notify(
                    title='ახალი ოსტატი',
                    message=f'ახალი ოსტატი: {lng}',
                    app_name='xeloba.ge',
                    timeout=10,
                    app_icon=image_path
                )
                print('notify ok')
        else:
            notification.notify(
                title='ახალი ოსტატი არ არის',
                message='',
                app_name='xeloba.ge',
                timeout=10,
                app_icon=image_path
            )
            print('notify unok')
    else:
        print(f'Failed to fetch data from API, Status Code: {response.status_code}')

# Example usage
# newUserNotify()

def index(request):
    c = {'csrf_token': get_token(request)} 
    template = loader.get_template('index.html')
    return HttpResponse(template.render(c, request))  
def reg(request):
    if request.method == "POST":
           
        user_data = {
            "userID": request.POST.get('userId'),
            "setDate": dateshorts,
            "name": request.POST.get('firstName'),
            "lastname": request.POST.get('lastName'),
            "personalID": request.POST.get('personalIdNumber'),
            "email": request.POST.get('email'),
            "phoneNumber": request.POST.get('phoneNumber'),
            "password": request.POST.get('password'),
            "city": request.POST.get('city'),
            "district": request.POST.get('district'),
            "status": request.POST.get('verificationStatus'),
            "profession": request.POST.get('profession'),
            "experience": request.POST.get('experience'),
            "price": request.POST.get('price'),
            "visitPrice": request.POST.get("minCallFee"), 
            "moreInfo": request.POST.get("moreInformation"), 
            "profileImgs": request.POST.get("profilePicture"), 
        }

        # Read existing data from the JSON file or initialize an empty list
        json_file_path = './data.json'
        if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
            with open(json_file_path, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
        else:
            data = []

        # Append new data to the existing data
        data.append(user_data)

        # Write the updated data back to the JSON file
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, indent=4, ensure_ascii=False)

        c = {'csrf_token': get_token(request)} 
        template = loader.get_template('response.html')
        return HttpResponse(template.render(c, request))  

def edit_user(request):
    if request.method == 'POST':
        formatted_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        userid = request.POST.get('userid')
        mob = request.POST.get('phoneNumber')
        # New user data from the request
        new_user_data = {
            "userID": userid,  # Use the correct variable name to get user ID
            "setDate": dateshorts,
            "name": request.POST.get('name'),
            "lastname": request.POST.get('surName'),
            "personalID": request.POST.get('personalIdNumber'),
            "email": request.POST.get('email'),
            "phoneNumber": request.POST.get('phoneNumber'),
            "password": request.POST.get('password'),
            "city": request.POST.get('city'),
            "district": request.POST.get('district'),
            "status": request.POST.get('verify'),
            "profession": request.POST.get('mainProfession'),
            "experience": request.POST.get('experienceInTheMainProfession'),
            "price": request.POST.get('priceOfTheMainProfession')
        }

        # Read existing data from the JSON file or initialize an empty list
        json_file_path = './data.json'
        if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
            with open(json_file_path, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)
        else:
            data = []

        # Find the user with the matching userid and update their data
        user_found = False
        for user in data:
            if user['userID'] == userid:
                for key, value in new_user_data.items():
                    if value is not None:
                        user[key] = value
                user_found = True
                break

        # If user was not found, optionally append new user data
        if not user_found:
            data.append(new_user_data)

        # Write the updated data back to the JSON file
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, indent=4, ensure_ascii=False)

        
        # Render the response with updated user data
        


        return render(request, 'adminpanel.html', {'users': data})



def delete_user(request):
    if request.method == "POST":
        try:
            # Get the user ID from the POST data
            delid = request.POST.get('userDel')
            ok = '''
                <div class="alert alert-warning alert-dismissible fade show" role="alert">
                    <strong>წაიშალა! მომხმარებელი </strong>
                    <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
                </div>
            '''
            if not delid:
                return JsonResponse({"error": "No user ID provided"}, status=400)

            delid = int(delid)

            # Read existing data from the JSON file
            json_file_path = './data.json'  # Ensure this path is correct relative to your project
            if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
                with open(json_file_path, 'r', encoding='utf-8') as json_file:
                    data = json.load(json_file)
            else:
                return JsonResponse({"error": "Data file not found or empty"}, status=404)

            # Filter out the user with the given user ID
            data = [user for user in data if user.get('userID') != delid]

            # Write the updated data back to the JSON file
            with open(json_file_path, 'w', encoding='utf-8') as json_file:
                json.dump(data, json_file, indent=4, ensure_ascii=False)

            return render(request, 'adminpanel.html', {'users': data, 'categories': categories,'smsbalance': bl} )
        except Exception as e:
            return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)
    else:
        return JsonResponse({"error": "Method not allowed"}, status=405)
# def send_sms(destination, content):
#     api_url = 'https://smsoffice.ge/api/v2/send/'
#     params = {
#         'key': 'f5e53ecd65894293934d40e5c3bcd901',  # Your API key
#         'destination': destination,  # The phone number to send the SMS to
#         'sender': 'Xeloba.ge',  # The sender name
#         'content': content,  # The SMS content
#         'urgent': 'true'  # Optional: mark the message as urgent
#     }

#     response = requests.post(api_url, data=params)

#     if response.status_code == 200:
#         print(f'SMS sent successfully to {destination}!')
#         print('Response:', response.json())
#     else:
#         print(f'Failed to send SMS to {destination}')
#         print('Status Code:', response.status_code)
#         print('Response:', response.text)

# def job():
#     destination = '568809304'
#     content = f'დღიური რეპორტი: ახალი 20 მომხმარებელი აქიდან 14 ხელოსანი, დაიხარჯა 18 სმს, დარჩა: {bl}'
#     print(bl)
#     send_sms(destination, content)

# # Schedule the job every 22 hours
# schedule.every(1).minutes.do(job)

# # Keep the script running to check the schedule
# while True:
#     schedule.run_pending()
#     smstimer.sleep(1)

def add_user(request):
    if request.method == "POST":
        try:
            user_data = Craftsman(
                user_id=rn,
                set_date=formatted_datetime,
                name=request.POST.get('name'),
                lastname=request.POST.get('surName'),
                personal_id=request.POST.get('personalID'),
                email=request.POST.get('email'),
                phone_number=request.POST.get('mobile'),
                password=request.POST.get('password'),
                city=request.POST.get('city'),
                district=request.POST.get('district'),
                profession=request.POST.get('profession'),
                experience=request.POST.get('experience'),
                salary=request.POST.get('salary'),
                last_login= formatted_datetime,
                image=request.FILES.get('profilimg')
            )
            user_data.save()

           

            return render(request, 'adminpanel.html', {'users': data, 'categories': categories,'smsbalance': bl} )
        except Exception as e:
            return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)
    else:
        return JsonResponse({"error": "Method not allowed"}, status=405)

def verify_user(request):
    if request.method == "POST":
        mob = request.POST.get('phoneNumber')
        name = request.POST.get('name')
        user_id = request.POST.get('userid')

        # Simulate response with status '1' for the purpose of this example
        response = {'status': '1'}  # Replace with actual logic to get the response status

        if response['status'] == '1':
            # Read existing data from the JSON file
            json_file_path = './data.json'
            with open(json_file_path, 'r', encoding='utf-8') as json_file:
                data = json.load(json_file)

            # Find the user with the matching user_id and update their status
            user_found = False
            for user in data:
                if user['userID'] == user_id:
                    user['status'] = "1"
                    user_found = True
                    break

            if user_found:
                # Write the updated data back to the JSON file
                with open(json_file_path, 'w', encoding='utf-8') as json_file:
                    json.dump(data, json_file, indent=4, ensure_ascii=False)

                # Send the SMS notification
                api_url = 'https://smsoffice.ge/api/v2/send/'
                params = {
                    'key': 'f5e53ecd65894293934d40e5c3bcd901',  # Your API key
                    'destination': mob,  # The phone number to send the SMS to
                    'sender': 'Xeloba.ge',  # The sender name
                    'content': f'{name} შენი პროფილი ({user_id}) გააქტიურდა',  # The SMS content
                    'urgent': 'true'  # Optional: mark the message as urgent
                }

                sms_response = requests.post(api_url, data=params)

                if sms_response.status_code == 200:
                    print('SMS sent successfully')
                    s = sms_response.json()
                    
                    user_data = {
                        "userID": user_id [
                            'smsDelivery': True,
                        ],
                    
                    }

                    # Read existing data from the JSON file or initialize an empty list
                    json_file_path = './smsfullreport.json'
                    if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
                        with open(json_file_path, 'r', encoding='utf-8') as json_file:
                            data = json.load(json_file)
                    else:
                        data = []

                    # Append new data to the existing data
                    data.append(user_data)

                    # Write the updated data back to the JSON file
                    with open(json_file_path, 'w', encoding='utf-8') as json_file:
                        json.dump(data, json_file, indent=4, ensure_ascii=False)




                else:
                    print('Failed to send SMS')
                    print('Status Code:', sms_response.status_code)
                    print('Response:', sms_response.text)
            else:
                print(f'User with ID {user_id} not found.')

    # Render the admin panel with necessary context
    return render(request, 'adminpanel.html', {
        'users': data, 
        'categories': categories,
        'smsbalance': bl
    })


# # sos sms 
# if bl > 50:

#     api_url = 'https://smsoffice.ge/api/v2/send/'
#     params = {
#         'key': 'f5e53ecd65894293934d40e5c3bcd901',  # Your API key
#         'destination': '571107460',  # The phone number to send the SMS to
#         'sender': 'Xeloba.ge',  # The sender name
#         'content': f'\U0001F198 ყურადღება! SMS ბალანსი ({bl}) ნაკლებია',  # The SMS content
#         'urgent': 'true'  # Optional: mark the message as urgent
#     }

#     sms_response = requests.post(api_url, data=params)

#     if sms_response.status_code == 200:
#         print('SMS sent successfully')
#         print('Response:', sms_response.json())
#     else:
#         print('Failed to send SMS')
#         print('Status Code:', sms_response.status_code)
#         print('Response:', sms_response.text)
